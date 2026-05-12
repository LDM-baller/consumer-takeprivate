"""Export Consumer take-privates xlsx -> JSON for the static web app.

Usage: python export.py [path-to-xlsx]
Default: ./source.xlsx
Output: ./data/deals.json, ./data/lulu.json, ./data/prices/{TICKER}.json
"""
import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

ROOT = Path(__file__).parent
SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "source.xlsx"
DATA = ROOT / "data"
PRICES = DATA / "prices"
DATA.mkdir(exist_ok=True)
PRICES.mkdir(exist_ok=True)

DEAL_TICKERS = ["WBA", "OLPX", "CPRI", "TAST", "CHS", "DSEY", "TWNK", "M", "ROVR", "SOVO"]


def iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    return v


def num_or_none(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def read_prices(ws):
    rows = []
    r = 16
    while True:
        d = ws.cell(r, 3).value
        p = ws.cell(r, 2).value
        if d is None and p is None:
            break
        if isinstance(d, datetime) and isinstance(p, (int, float)):
            rows.append([d.date().isoformat(), float(p)])
        r += 1
    return rows


def main():
    wb = load_workbook(SRC, data_only=True)

    # --- Summary -> deals.json ---
    s = wb["Summary"]
    header_row = 4
    headers = [s.cell(header_row, c).value for c in range(1, s.max_column + 1)]
    col = {h: i for i, h in enumerate(headers)}

    def get(row, name):
        idx = col.get(name)
        return row[idx] if idx is not None else None

    deals = []
    for r in range(header_row + 1, s.max_row + 1):
        row = [s.cell(r, c).value for c in range(1, s.max_column + 1)]
        ticker = get(row, "Ticker")
        if not ticker:
            continue
        deals.append({
            "ticker": ticker,
            "company": get(row, "Company"),
            "acquirer": get(row, "Acquirer"),
            "acquirerType": get(row, "Acquirer Type"),
            "outcome": get(row, "Outcome"),
            "outcomeReason": get(row, "Outcome Reason"),
            "description": get(row, "Target Description"),
            "context": get(row, "Deal Context"),
            "takeoutPrice": num_or_none(get(row, "Takeout Price")),
            "unaffectedDate": iso(get(row, "Unaffected Date")),
            "announceDate": iso(get(row, "Announce Date")),
            "endDatePulled": iso(get(row, "End-Date Pulled")),
            "unaffectedClose": num_or_none(get(row, "Unaffected Close")),
            "avg12mo": num_or_none(get(row, "12mo Avg")),
            "avg24mo": num_or_none(get(row, "24mo Avg")),
            "high52wk": num_or_none(get(row, "52-Week High")),
            "premUnaff": num_or_none(get(row, "Premium to Unaffected")),
            "prem12mo": num_or_none(get(row, "Premium to 12mo Avg")),
            "prem24mo": num_or_none(get(row, "Premium to 24mo Avg")),
            "prem52wk": num_or_none(get(row, "Premium to 52wk High")),
        })

    deals.sort(key=lambda d: d["announceDate"] or "", reverse=True)
    (DATA / "deals.json").write_text(json.dumps(deals, indent=2))
    print(f"Wrote deals.json ({len(deals)} deals)")

    # --- LULU -> lulu.json ---
    lulu_ws = wb["LULU"]
    lulu = {
        "ticker": "LULU",
        "company": "Lululemon Athletica",
        "anchorDate": iso(lulu_ws.cell(5, 2).value),
        "latestClose": num_or_none(lulu_ws.cell(3, 7).value),
        "avg12mo": num_or_none(lulu_ws.cell(4, 7).value),
        "avg24mo": num_or_none(lulu_ws.cell(5, 7).value),
        "high52wk": num_or_none(lulu_ws.cell(12, 7).value),
        "dilutedShares": 117.0,
        "netCash": 1800.0,
    }
    (DATA / "lulu.json").write_text(json.dumps(lulu, indent=2))
    print(f"Wrote lulu.json: latest=${lulu['latestClose']:.2f}")

    # --- Price history per ticker ---
    for tk in DEAL_TICKERS + ["LULU"]:
        if tk not in wb.sheetnames:
            print(f"  WARN: missing tab {tk}")
            continue
        prices = read_prices(wb[tk])
        (PRICES / f"{tk}.json").write_text(json.dumps(prices))
        print(f"  {tk}: {len(prices)} points")


if __name__ == "__main__":
    main()
